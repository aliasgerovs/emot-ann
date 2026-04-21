"""
comparison.py
-------------
Reliability coding comparison tool for Student C.
Run from VS Code terminal: python comparison.py
Opens automatically in Chrome at http://localhost:7862

Workflow:
  Step 1 — Upload two CSV files → download comparison Excel
            (agreed cells auto-filled in Final Code column, disagreements left blank/yellow)
  Step 2 — Fill in yellow cells in Excel, re-upload → download final CSV
"""

import io
import gradio as gr
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Behavior columns (must match regulation.py output) ──────────────────────
BEHAVIOR_COLS = [
    "NC", "FocusedOnStimulus", "Distraction", "InteractionWithFamily",
    "InteractionWithExperimenter", "SelfComforting", "Persistence",
    "CheckingInWithAdults", "OutOfChair", "PresenceOfOthersInRoom"
]
META_COLS = ["ClipNumber", "StartTime", "EndTime"]

# ── Colors ───────────────────────────────────────────────────────────────────
RED_FILL     = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
GREEN_FILL   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
YELLOW_FILL  = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
GREY_FILL    = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
BLUE_FILL    = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
FINAL_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
FINAL_HDR    = PatternFill("solid", start_color="375623", end_color="375623")
NAVY_FILL    = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
MISSING_FILL = PatternFill("solid", start_color="F2CEEF", end_color="F2CEEF")
ORANGE_FILL  = PatternFill("solid", start_color="F4B942", end_color="F4B942")

THIN   = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def bold_cell(cell, fill=None, font_color="000000"):
    cell.font      = Font(bold=True, color=font_color, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        cell.fill = fill
    cell.border = BORDER


def normal_cell(cell, fill=None, italic=False):
    cell.font      = Font(name="Arial", size=10, italic=italic)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        cell.fill = fill
    cell.border = BORDER


def load_csv(path):
    df = pd.read_csv(path)
    df["ClipNumber"] = pd.to_numeric(df["ClipNumber"], errors="coerce")
    return df.reset_index(drop=True)


def build_clip_index(df):
    return {int(row["ClipNumber"]): row for _, row in df.iterrows()}


# ── Step 1: Build comparison Excel ──────────────────────────────────────────
def build_comparison_excel(df1, df2):
    name1       = df1["AnnotatorName"].iloc[0] if "AnnotatorName" in df1.columns else "Coder 1"
    name2       = df2["AnnotatorName"].iloc[0] if "AnnotatorName" in df2.columns else "Coder 2"
    participant = df1["ParticipantID"].iloc[0]  if "ParticipantID" in df1.columns else "?"
    task        = df1["TaskType"].iloc[0]       if "TaskType"      in df1.columns else "?"

    clips1       = build_clip_index(df1)
    clips2       = build_clip_index(df2)
    all_clips    = sorted(set(clips1) | set(clips2))
    only_in_1    = sorted(set(clips1) - set(clips2))
    only_in_2    = sorted(set(clips2) - set(clips1))
    common_clips = sorted(set(clips1) & set(clips2))

    beh_count  = len(BEHAVIOR_COLS)
    meta_count = len(META_COLS)

    c1_start    = meta_count + 1
    c2_start    = c1_start + beh_count
    match_start = c2_start + beh_count
    final_start = match_start + 3
    last_col    = final_start + beh_count - 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    tc = ws["A1"]
    tc.value     = f"Participant: {participant}  |  Task: {task}  |  {name1}  vs  {name2}"
    tc.font      = Font(bold=True, size=12, name="Arial", color="FFFFFF")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.fill      = NAVY_FILL
    ws.row_dimensions[1].height = 24

    # Row 2: warning banner (if needed)
    warnings = []
    if only_in_1:
        warnings.append(f"Clips only in {name1}: {only_in_1}")
    if only_in_2:
        warnings.append(f"Clips only in {name2}: {only_in_2}")
    if len(df1) != len(df2):
        warnings.append(f"Row count differs ({name1}: {len(df1)}, {name2}: {len(df2)}) — task may have ended early")

    if warnings:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        wc = ws["A2"]
        wc.value     = "⚠️  " + "   |   ".join(warnings)
        wc.font      = Font(bold=True, size=10, name="Arial")
        wc.alignment = Alignment(horizontal="left", vertical="center")
        wc.fill      = ORANGE_FILL
        wc.border    = BORDER
        ws.row_dimensions[2].height = 20
        header_row  = 3
        subhead_row = 4
        data_start  = 5
    else:
        header_row  = 2
        subhead_row = 3
        data_start  = 4

    # Group headers
    ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=meta_count)
    bold_cell(ws.cell(header_row, 1, "Meta"), GREY_FILL)

    ws.merge_cells(start_row=header_row, start_column=c1_start, end_row=header_row, end_column=c1_start+beh_count-1)
    bold_cell(ws.cell(header_row, c1_start, name1), YELLOW_FILL)

    ws.merge_cells(start_row=header_row, start_column=c2_start, end_row=header_row, end_column=c2_start+beh_count-1)
    bold_cell(ws.cell(header_row, c2_start, name2), GREEN_FILL)

    ws.merge_cells(start_row=header_row, start_column=match_start, end_row=header_row, end_column=match_start+1)
    bold_cell(ws.cell(header_row, match_start, "Agreement"), GREY_FILL)

    ws.merge_cells(start_row=header_row, start_column=final_start, end_row=header_row, end_column=last_col)
    bold_cell(ws.cell(header_row, final_start, "✏️  FINAL CODE  —  fill in yellow cells only"), FINAL_HDR, "FFFFFF")

    # Column labels
    for i, col in enumerate(META_COLS, start=1):
        bold_cell(ws.cell(subhead_row, i, col), GREY_FILL)
        ws.column_dimensions[get_column_letter(i)].width = 9

    for i, col in enumerate(BEHAVIOR_COLS, start=c1_start):
        bold_cell(ws.cell(subhead_row, i, col[:6]), YELLOW_FILL)
        ws.column_dimensions[get_column_letter(i)].width = 7

    for i, col in enumerate(BEHAVIOR_COLS, start=c2_start):
        bold_cell(ws.cell(subhead_row, i, col[:6]), GREEN_FILL)
        ws.column_dimensions[get_column_letter(i)].width = 7

    bold_cell(ws.cell(subhead_row, match_start,   "Agree"), GREY_FILL)
    bold_cell(ws.cell(subhead_row, match_start+1, "%"),     GREY_FILL)
    ws.column_dimensions[get_column_letter(match_start)].width   = 8
    ws.column_dimensions[get_column_letter(match_start+1)].width = 7
    ws.column_dimensions[get_column_letter(match_start+2)].width = 2

    for i, col in enumerate(BEHAVIOR_COLS, start=final_start):
        bold_cell(ws.cell(subhead_row, i, col[:6]), FINAL_HDR, "FFFFFF")
        ws.column_dimensions[get_column_letter(i)].width = 7

    ws.row_dimensions[subhead_row].height = 28

    # Data rows
    total_agree   = 0
    total_cells   = 0
    missing_clips = []

    for r, clip in enumerate(all_clips):
        row_excel = data_start + r
        row1_data = clips1.get(clip)
        row2_data = clips2.get(clip)

        missing_in = name1 if row1_data is None else (name2 if row2_data is None else None)
        meta_src   = row1_data if row1_data is not None else row2_data

        for i, col in enumerate(META_COLS, start=1):
            val = meta_src.get(col, "") if meta_src is not None else ""
            normal_cell(ws.cell(row_excel, i, val), MISSING_FILL if missing_in else GREY_FILL)

        if missing_in:
            missing_clips.append((clip, missing_in))
            # Show the available coder's values; mark the missing coder's columns purple
            for j, col in enumerate(BEHAVIOR_COLS):
                # Coder 1 column
                c1_cell = ws.cell(row_excel, c1_start + j)
                if row1_data is not None and missing_in != name1:
                    c1_cell.value = int(row1_data[col]) if col in row1_data.index else None
                    normal_cell(c1_cell)
                else:
                    c1_cell.value = None
                    normal_cell(c1_cell, MISSING_FILL)

                # Coder 2 column
                c2_cell = ws.cell(row_excel, c2_start + j)
                if row2_data is not None and missing_in != name2:
                    c2_cell.value = int(row2_data[col]) if col in row2_data.index else None
                    normal_cell(c2_cell)
                else:
                    c2_cell.value = None
                    normal_cell(c2_cell, MISSING_FILL)

                # Final Code — always yellow (decided in meeting)
                normal_cell(ws.cell(row_excel, final_start + j), YELLOW_FILL)

            # Agreement columns — N/A for missing rows
            normal_cell(ws.cell(row_excel, match_start,   "—"), GREY_FILL)
            normal_cell(ws.cell(row_excel, match_start+1, "—"), GREY_FILL)
            continue

        agree_count = 0
        for j, col in enumerate(BEHAVIOR_COLS):
            v1 = int(row1_data[col]) if col in row1_data.index else None
            v2 = int(row2_data[col]) if col in row2_data.index else None

            normal_cell(ws.cell(row_excel, c1_start + j, v1),
                        GREEN_FILL if (v1 is not None and v1 == v2) else RED_FILL)
            normal_cell(ws.cell(row_excel, c2_start + j, v2),
                        GREEN_FILL if (v2 is not None and v1 == v2) else RED_FILL)

            fc = ws.cell(row_excel, final_start + j)
            if v1 is not None and v2 is not None and v1 == v2:
                fc.value = v1
                normal_cell(fc, FINAL_FILL)
                agree_count += 1
            else:
                fc.value = None
                normal_cell(fc, YELLOW_FILL)

        total_agree += agree_count
        total_cells += beh_count

        pct  = round(agree_count / beh_count * 100, 1)
        pfil = GREEN_FILL if pct == 100 else (RED_FILL if pct < 80 else None)
        normal_cell(ws.cell(row_excel, match_start,   agree_count), pfil)
        normal_cell(ws.cell(row_excel, match_start+1, f"{pct}%"),   pfil)

    # Summary row
    sum_row     = data_start + len(all_clips)
    overall_pct = round(total_agree / total_cells * 100, 1) if total_cells else 0
    ws.cell(sum_row, 1, "TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(sum_row, match_start,   total_agree).font   = Font(bold=True, name="Arial")
    ws.cell(sum_row, match_start+1, f"{overall_pct}%").font = Font(bold=True, name="Arial")
    for ci in range(1, last_col + 1):
        ws.cell(sum_row, ci).fill   = GREY_FILL
        ws.cell(sum_row, ci).border = BORDER

    ws.freeze_panes = f"A{data_start}"

    # Sheet 2: Behavior Summary
    ws2 = wb.create_sheet("Behavior Summary")
    for ci, h in enumerate(["Behavior", "Clips Compared", "Agreed", "Disagreed", "Agreement %"], 1):
        bold_cell(ws2.cell(1, ci, h), BLUE_FILL, "000000")
        ws2.column_dimensions[get_column_letter(ci)].width = 20

    for ri, col in enumerate(BEHAVIOR_COLS, start=2):
        agree = disagree = 0
        for clip in common_clips:
            v1 = int(clips1[clip][col]) if col in clips1[clip].index else None
            v2 = int(clips2[clip][col]) if col in clips2[clip].index else None
            if v1 is not None and v2 is not None:
                if v1 == v2: agree += 1
                else:        disagree += 1
        total = agree + disagree
        pct   = round(agree / total * 100, 1) if total else 0
        fill  = GREEN_FILL if pct >= 80 else RED_FILL
        for ci, val in enumerate([col, total, agree, disagree, f"{pct}%"], 1):
            normal_cell(ws2.cell(ri, ci, val), fill)

    if missing_clips:
        start_r = len(BEHAVIOR_COLS) + 3
        ws2.merge_cells(start_row=start_r, start_column=1, end_row=start_r, end_column=3)
        bold_cell(ws2.cell(start_r, 1, "Missing Clips"), ORANGE_FILL)
        for ri2, (clip, who) in enumerate(missing_clips, start=start_r+1):
            ws2.cell(ri2, 1, clip).font = Font(name="Arial", size=10)
            ws2.cell(ri2, 2, f"Missing in {who}").font = Font(name="Arial", size=10, italic=True)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    summary = {
        "participant": participant,
        "task": task,
        "name1": name1,
        "name2": name2,
        "total_clips": len(all_clips),
        "compared": len(common_clips),
        "missing": len(missing_clips),
        "overall_pct": overall_pct,
        "yellow_cells": total_cells - total_agree,
        "missing_clips": missing_clips,
    }
    return buf.getvalue(), summary


# ── Step 2: Export final CSV from filled Excel ───────────────────────────────
MISSING_RGB = "FFF2CEEF"

def get_rgb(cell):
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
            return cell.fill.fgColor.rgb
    except Exception:
        pass
    return None


def export_final_csv(xlsx_bytes):
    from openpyxl.cell.cell import MergedCell

    wb = load_workbook(io.BytesIO(xlsx_bytes))

    if "Comparison" not in wb.sheetnames:
        return None, "❌ 'Comparison' sheet not found. Make sure you uploaded the correct file."

    ws = wb["Comparison"]

    # Find Final Code start column
    final_start_col = None
    for check_row in [2, 3]:
        for cell in ws[check_row]:
            if isinstance(cell, MergedCell):
                continue
            if cell.value and "FINAL CODE" in str(cell.value):
                final_start_col = cell.column
                break
        if final_start_col:
            break

    if final_start_col is None:
        return None, "❌ Could not find the FINAL CODE column. Is this the right file?"

    # Collect data rows
    data_rows    = []
    missing_rows = []

    for row in ws.iter_rows(min_row=2):
        first_val = row[0].value
        if str(first_val) == "TOTAL":
            break
        try:
            clip_num = int(first_val)
        except (TypeError, ValueError):
            continue

        # A MISSING row has purple fill in the first coder column
        # (no longer uses MergedCells - both coders have separate columns now)
        first_coder_col = final_start_col - 2 * len(BEHAVIOR_COLS) - 3  # c1_start approx
        # Simpler: check if BOTH coder columns for first behavior are empty (purple)
        # We detect missing rows by checking if the row has purple fill on any cell
        # Most reliable: check if both coder 1 and coder 2 first-behavior cells are empty
        # Actually, just treat ALL rows as data — empty yellow cells will be flagged below
        data_rows.append(row)

    # Check for empty yellow cells
    empty_cells = []
    for row in data_rows:
        row_idx = row[0].row
        clip    = row[0].value
        for j, col in enumerate(BEHAVIOR_COLS):
            cell = ws.cell(row_idx, final_start_col + j)
            if cell.value is None or str(cell.value).strip() == "":
                empty_cells.append(f"Clip {clip} → {col}")

    if empty_cells:
        msg = (
            f"⚠️ **{len(empty_cells)} yellow cell(s) are still empty.**\n\n"
            + "\n".join(f"- {e}" for e in empty_cells)
            + "\n\nPlease fill them in and re-upload the file."
        )
        return None, msg

    # Extract participant and task from title
    title_val   = ws["A1"].value or ""
    participant = task = "?"
    try:
        participant = title_val.split("Participant:")[1].split("|")[0].strip()
        task        = title_val.split("Task:")[1].split("|")[0].strip()
    except Exception:
        pass

    records = []
    for row in data_rows:
        row_idx = row[0].row
        record  = {
            "ParticipantID": participant,
            "TaskType":      task,
            "AnnotatorName": "Final",
            "ClipNumber":    ws.cell(row_idx, 1).value,
            "StartTime":     ws.cell(row_idx, 2).value,
            "EndTime":       ws.cell(row_idx, 3).value,
        }
        nc_val = ws.cell(row_idx, final_start_col).value  # NC is first behavior col
        is_nc  = (int(nc_val) == 1) if nc_val is not None else False
        for j, col in enumerate(BEHAVIOR_COLS):
            if is_nc and col != "NC":
                record[col] = ""  # NC=1: leave all other behaviors blank
            else:
                val = ws.cell(row_idx, final_start_col + j).value
                record[col] = int(val) if val is not None else ""
        records.append(record)

    df_out  = pd.DataFrame(records)
    csv_buf = io.StringIO()
    df_out.to_csv(csv_buf, index=False)

    msg = f"✅ **Final CSV ready!** ({len(df_out)} rows)"
    if missing_rows:
        msg += f"\n\nℹ️ Missing clips not included: {missing_rows}"

    return csv_buf.getvalue(), msg


# ── Gradio UI ────────────────────────────────────────────────────────────────
css = """
#compare_btn, #export_btn {
    background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    border: none;
    color: white;
    font-weight: 600;
}
.step-header {
    background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 15px;
    border-radius: 10px;
    margin: 20px 0 10px 0;
    font-size: 18px;
    font-weight: 600;
    text-align: center;
}
"""

# Store Excel bytes between steps
_excel_store = {"data": None, "filename": None}

with gr.Blocks(css=css, title="Reliability Comparison Tool") as demo:
    gr.HTML("""
        <div style='text-align: center; padding: 20px;'>
            <h1>🔍 Reliability Comparison Tool</h1>
            <p>Upload two coding CSV files to generate a comparison sheet, then export the final codes.</p>
        </div>
    """)

    gr.HTML("<div style='max-width: 1200px; margin: 0 auto;'>")

    # ── Step 1 ───────────────────────────────────────────────────────────────
    gr.HTML("<div class='step-header'>📂 Step 1: Upload Coding Files & Download Comparison</div>")

    with gr.Row():
        csv1_input = gr.File(label="Coder 1 CSV", file_types=[".csv"])
        csv2_input = gr.File(label="Coder 2 CSV", file_types=[".csv"])

    compare_btn    = gr.Button("🔍 Generate Comparison", elem_id="compare_btn", size="lg")
    compare_status = gr.Markdown("")
    excel_download = gr.File(label="⬇️ Download Comparison Excel", visible=False)

    # ── Step 2 ───────────────────────────────────────────────────────────────
    gr.HTML("<div class='step-header'>✏️ Step 2: Fill Yellow Cells in Excel, Then Upload & Export Final CSV</div>")

    gr.Markdown(
        "1. Open the downloaded Excel file\n"
        "2. In the **FINAL CODE** section, fill in the **yellow cells** with the agreed values (0 or 1)\n"
        "3. Save the file and upload it below"
    )

    excel_upload  = gr.File(label="Upload Filled Excel", file_types=[".xlsx"])
    export_btn    = gr.Button("📤 Export Final CSV", elem_id="export_btn", size="lg")
    export_status = gr.Markdown("")
    csv_download  = gr.File(label="⬇️ Download Final CSV", visible=False)

    gr.HTML("</div>")

    # ── Handlers ─────────────────────────────────────────────────────────────
    def handle_compare(csv1_file, csv2_file):
        if csv1_file is None or csv2_file is None:
            return "⚠️ Please upload both CSV files.", gr.update(visible=False)

        try:
            df1 = load_csv(csv1_file.name)
            df2 = load_csv(csv2_file.name)
        except Exception as e:
            return f"❌ Failed to read CSV files: {e}", gr.update(visible=False)

        try:
            excel_bytes, summary = build_comparison_excel(df1, df2)
        except Exception as e:
            return f"❌ Failed to build comparison: {e}", gr.update(visible=False)

        # Save to a temp file Gradio can serve
        import tempfile, os
        fname = f"{summary['participant']}_Regulation_{summary['task']}_{summary['name1']}_{summary['name2']}_Comparison.xlsx"
        tmp_dir = tempfile.mkdtemp()
        tmp_path = os.path.join(tmp_dir, fname)
        with open(tmp_path, "wb") as tmp:
            tmp.write(excel_bytes)

        missing_note = ""
        if summary["missing_clips"]:
            lines = [f"  - Clip {c} → missing in {w}" for c, w in summary["missing_clips"]]
            missing_note = "\n\n🟣 **Missing clips (marked purple in Excel):**\n" + "\n".join(lines)

        warn_note = ""
        if summary["missing"] > 0 or summary["total_clips"] != summary["compared"] + summary["missing"]:
            pass  # already covered by missing_note

        msg = (
            f"✅ **Comparison ready!**\n\n"
            f"**Participant:** {summary['participant']}  |  **Task:** {summary['task']}\n\n"
            f"**Coders:** {summary['name1']}  vs  {summary['name2']}\n\n"
            f"**Clips compared:** {summary['compared']} / {summary['total_clips']}  |  "
            f"**Missing:** {summary['missing']}  |  "
            f"**Overall agreement:** {summary['overall_pct']}%\n\n"
            f"**Yellow cells to fill in:** {summary['yellow_cells']}"
            + missing_note
        )

        return msg, gr.update(value=tmp_path, visible=True, label=f"⬇️ {fname}")

    def handle_export(excel_file):
        if excel_file is None:
            return "⚠️ Please upload the filled Excel file.", gr.update(visible=False)

        try:
            with open(excel_file.name, "rb") as f:
                xlsx_bytes = f.read()
        except Exception as e:
            return f"❌ Could not read file: {e}", gr.update(visible=False)

        csv_str, msg = export_final_csv(xlsx_bytes)

        if csv_str is None:
            return msg, gr.update(visible=False)

        import tempfile
        # Derive filename from participant/task extracted in export_final_csv
        # Re-read title from the xlsx to get participant and task
        import io as _io
        from openpyxl import load_workbook as _lw
        _wb = _lw(_io.BytesIO(xlsx_bytes))
        _ws = _wb["Comparison"]
        _title = _ws["A1"].value or ""
        try:
            _participant = _title.split("Participant:")[1].split("|")[0].strip()
            _task        = _title.split("Task:")[1].split("|")[0].strip()
        except Exception:
            _participant = "Unknown"
            _task        = "Unknown"
        fname = f"{_participant}_Regulation_{_task}_Final.csv"

        tmp_dir2 = tempfile.mkdtemp()
        tmp_path2 = os.path.join(tmp_dir2, fname)
        with open(tmp_path2, "wb") as tmp:
            tmp.write(csv_str.encode("utf-8"))

        return msg, gr.update(value=tmp_path2, visible=True, label=f"⬇️ {fname}")

    import os

    compare_btn.click(
        handle_compare,
        inputs=[csv1_input, csv2_input],
        outputs=[compare_status, excel_download]
    )

    export_btn.click(
        handle_export,
        inputs=[excel_upload],
        outputs=[export_status, csv_download]
    )


if __name__ == "__main__":
    port = 7862

    print("\n" + "="*60)
    print("🔍 RELIABILITY COMPARISON TOOL")
    print("="*60)
    print(f"Local URL: http://localhost:{port}")
    print("="*60 + "\n")

    demo.queue(max_size=10, api_open=False)
    demo.launch(
        share=False,
        server_port=port,
        inbrowser=True,
        show_error=True,
        quiet=False
    )
